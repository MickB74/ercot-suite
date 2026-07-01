"""Render the "This month & next — weather forecast" tab in each portal.

Each portal calls :func:`render_near_term_tab` after setting up its hub, analytics,
contract, and branding imports. The function is self-contained: it fetches weather,
calibrates against SCED, merges actual + forecast daily data, and renders the chart
and KPIs.

Usage::

    from ercot_core.near_term_bill import render_near_term_tab

    render_near_term_tab(
        st=st,
        a=contract.ASSET,
        hub=hub,
        analytics=analytics,
        branding=branding,
        terms=terms,
        win_start=win_start,
        win_end=win_end,
        hist_mwh=hist_mwh,          # pd.Series indexed 1-12 (historical shape)
        fwd_price=fwd_price,         # forward market price $/MWh from the sidebar
        gen_kwargs=None,             # extra kwargs for hub.generation() (Azure Sky multi-unit)
    )
"""

from __future__ import annotations

import datetime as dt

import pandas as pd
import plotly.graph_objects as go

from . import gen_forecast as gf
from . import weather_forecast as wf


def _hub_ws_series(df: pd.DataFrame, hub_h: float):
    """Hub-height wind speed (m/s) from whichever level a frame carries, or None.

    Mirrors :func:`gen_forecast._wind_hourly_mw`'s level preference (120→100→80)
    so the forecast and the ERA5 archive are compared at the same height even
    though they expose different native levels.
    """
    for col, h in (("wind_speed_120m", 120.0), ("wind_speed_100m", 100.0),
                   ("wind_speed_80m", 80.0)):
        if col in df.columns:
            s = pd.to_numeric(df[col], errors="coerce")
            if s.fillna(0.0).sum() > 0:
                return s * (hub_h / h) ** (1.0 / 7.0)
    return None


def _wind_bias_ratio(forecast_df: pd.DataFrame, archive_df, hub_h: float):
    """ERA5/forecast mean hub-wind ratio over their overlap (None if not comparable).

    The forecast API's winds run hotter than the ERA5 archive that ``cal_factor``
    is fit against; scaling forecast winds by this ratio removes the cross-product
    bias so the calibrated model doesn't over-predict the forward forecast.
    """
    if archive_df is None or getattr(archive_df, "empty", True):
        return None
    f = _hub_ws_series(forecast_df, hub_h)
    a = _hub_ws_series(archive_df, hub_h)
    if f is None or a is None:
        return None
    j = pd.concat([f.rename("f"), a.rename("a")], axis=1).dropna()
    if len(j) < 48 or j["f"].mean() <= 0:   # need ~2 days of overlap
        return None
    return float(j["a"].mean() / j["f"].mean())


def _guard_forecast_months(forecast_rows, *, hist_mwh, cap_share, tech, strike,
                           fpx, blocked, fwin_start, fwin_end):
    """Clamp implausible forecast months to the historical monthly shape.

    Weather-driven months can imply impossible capacity factors: a hot GEFS run
    over-predicts (70%+ CF) and a short/under-covered climatological tail
    under-predicts (near-0%). For any forecast month whose implied CF falls
    outside a sane band — or that is materially under-covered — rebuild or rescale
    its unsettled days from ``hist_mwh`` (the historical monthly shape, already at
    the contracted share). Returns ``(rows, notes)`` where notes lists
    ``(ym, action, observed_cf)`` for anything the guard changed.
    """
    import calendar as _cal
    if hist_mwh is None or len(hist_mwh) == 0 or not cap_share or cap_share != cap_share:
        return forecast_rows, []
    cf_lo, cf_hi = (0.10, 0.60) if tech == "wind" else (0.06, 0.42)
    by_ym: dict[str, list] = {}
    for r in forecast_rows:
        by_ym.setdefault(r["date"].strftime("%Y-%m"), []).append(r)
    out, notes = [], []
    for ym, rows in by_ym.items():
        yr, mo = int(ym[:4]), int(ym[5:7])
        dim = _cal.monthrange(yr, mo)[1]
        intended = [dt.date(yr, mo, i + 1) for i in range(dim)]
        intended = [d for d in intended if fwin_start <= d <= fwin_end and d not in blocked]
        hist_month = float(hist_mwh.get(mo, float("nan")))
        fmwh = sum(x["mwh"] for x in rows)
        cf = fmwh / (cap_share * max(len(rows), 1) * 24.0)
        undercov = len(rows) < 0.9 * max(len(intended), 1)
        bad = undercov or not (cf_lo <= cf <= cf_hi)
        if not intended or hist_month != hist_month or hist_month <= 0 or not bad:
            out.extend(rows)                       # month is fine — leave untouched
            continue
        if undercov or fmwh <= 0:                  # coverage gap → fill every day flat
            daily = hist_month / dim
            for d in intended:
                out.append({"date": d, "mwh": daily, "net": daily * (fpx(d) - strike),
                            "price": fpx(d), "kind": f"histguard_{ym}"})
            notes.append((ym, "filled", round(cf, 3)))
        else:                                      # covered but impossible CF → rescale
            target = hist_month * (len(rows) / dim)
            scale = target / fmwh
            for x in rows:
                m2 = x["mwh"] * scale
                out.append({**x, "mwh": m2, "net": m2 * (fpx(x["date"]) - strike),
                            "kind": str(x["kind"]) + "_capped"})
            notes.append((ym, "rescaled", round(cf, 3)))
    out.sort(key=lambda r: r["date"])
    return out, notes


def render_near_term_tab(
    st,
    *,
    a: dict,
    hub,
    analytics,
    branding,
    terms: dict,
    win_start,
    win_end,
    hist_mwh: pd.Series,
    fwd_price: float,
    fwd_price_by_month: dict | None = None,
    fwd_band_df: pd.DataFrame | None = None,
    gen_kwargs: dict | None = None,
) -> None:
    """Render the near-term weather bill forecast tab.

    Parameters
    ----------
    st:
        The Streamlit module (passed in so this module doesn't import it at the
        top level, keeping it usable outside a Streamlit context).
    a:
        Portal ASSET dict (must have lat, lon, tech, capacity_mw; wind also
        needs hub_height_m).
    hub:
        Portal-specific hub module.
    analytics:
        Portal-specific analytics module.
    branding:
        Portal-specific branding module.
    terms:
        Loaded contract terms dict.
    win_start, win_end:
        Settlement data window (dates).
    hist_mwh:
        Historical mean monthly MWh indexed 1–12, at the contracted share.
        Used as fallback for days beyond the weather forecast horizon.
    fwd_price:
        Forward market price assumption ($/MWh) — usually set in the sidebar.
    gen_kwargs:
        Extra kwargs forwarded to ``hub.generation()`` — only needed for
        Azure Sky (``{"units": a["units"]}``).
    """
    strike = float(terms.get("strike", 0.0))
    share = float(terms.get("volume_share_pct", 100.0)) / 100.0
    cap_share = float(a["capacity_mw"]) * share
    is_wind = "wind" in str(a.get("tech", "")).lower()
    tech = "wind" if is_wind else "solar"
    hub_h = float(a.get("hub_height_m") or 90.0)
    cut_in_ms = float(a.get("cut_in_ms") or 3.0)
    rated_ms = float(a.get("rated_ms") or 12.0)
    cut_out_ms = float(a.get("cut_out_ms") or 25.0)
    # Real turbine model → the validated per-model power curve (same family
    # plant_value uses). None ⇒ gen_forecast falls back to the generic ramp.
    turbine_type = str(a.get("turbine_model") or "").strip() or None
    gen_kwargs = gen_kwargs or {}

    today_ct = pd.Timestamp.now("America/Chicago")
    month_start_ct = today_ct.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_month_start = month_start_ct + pd.offsets.MonthBegin(1)
    next_month_end_date = (next_month_start + pd.offsets.MonthEnd(1)).date()
    cur_month_str = month_start_ct.strftime("%Y-%m")
    next_month_str = next_month_start.strftime("%Y-%m")
    third_month_start = next_month_start + pd.offsets.MonthBegin(1)
    third_month_end_date = (third_month_start + pd.offsets.MonthEnd(1)).date()
    third_month_str = third_month_start.strftime("%Y-%m")
    fourth_month_start = third_month_start + pd.offsets.MonthBegin(1)
    fourth_month_end_date = (fourth_month_start + pd.offsets.MonthEnd(1)).date()
    fourth_month_str = fourth_month_start.strftime("%Y-%m")

    # ── fetch weather forecast (standard 16-day + GEFS ensemble 35-day) ─────
    @st.cache_data(show_spinner="Fetching weather forecast…", ttl=7200)
    def _weather(lat, lon, tech_key):
        try:
            return wf.fetch(lat, lon, tech_key, past_days=31, forecast_days=16), None
        except Exception as exc:  # noqa: BLE001
            return None, str(exc)

    @st.cache_data(show_spinner="Fetching medium-range forecast (GEFS 35-day)…", ttl=21600)
    def _medium_range(lat, lon, tech_key):
        try:
            return wf.fetch_medium_range(lat, lon, tech_key, forecast_days=35), None
        except Exception as exc:  # noqa: BLE001
            return None, str(exc)

    weather_df, wfail = _weather(float(a["lat"]), float(a["lon"]), tech)

    if wfail:
        st.warning(
            f"⚠️ Weather forecast unavailable ({wfail}). "
            "Check your internet connection or try again shortly."
        )
        return

    # Extend with GEFS ensemble beyond the 16-day standard horizon
    med_df, _ = _medium_range(float(a["lat"]), float(a["lon"]), tech)
    if med_df is not None and not weather_df.empty:
        std_end = weather_df.index[-1]
        ext = med_df[med_df.index > std_end]
        if not ext.empty:
            shared_cols = weather_df.columns.intersection(ext.columns)
            weather_df = pd.concat([weather_df, ext[shared_cols]])

    # Prior-year ERA5 for days beyond the GEFS horizon (~35 days).
    # Fetches the same calendar window from last year as a climatological proxy
    # — gives realistic day-to-day variation vs a flat monthly average.
    _clim_start = month_start_ct.date().replace(year=month_start_ct.year - 1)
    try:
        _clim_end = fourth_month_end_date.replace(year=fourth_month_end_date.year - 1)
    except ValueError:  # Feb 29 -> non-leap prior year
        _clim_end = fourth_month_end_date.replace(year=fourth_month_end_date.year - 1, day=28)

    @st.cache_data(show_spinner="Loading prior-year ERA5 (climatological baseline)…", ttl=86400)
    def _prior_year(lat, lon, tech_key, start_str, end_str):
        try:
            return wf.fetch_archive(lat, lon, tech_key, start_str, end_str), None
        except Exception as exc:  # noqa: BLE001
            return None, str(exc)

    py_df, _ = _prior_year(
        float(a["lat"]), float(a["lon"]), tech,
        str(_clim_start), str(_clim_end),
    )

    # ── calibrate against SCED using archive API ─────────────────────────────
    # The forecast endpoint's past_days returns 0 for shortwave_radiation
    # beyond ~30 days. SCED lags ~60 days, so we use the ERA5 archive endpoint
    # which provides accurate historical radiation for any date range.
    _units = list(gen_kwargs.get("units", []))  # empty list = single-unit portal
    _is_multi_unit = bool(_units)

    win_end_date = win_end if isinstance(win_end, dt.date) else pd.Timestamp(win_end).date()
    win_start_date = win_start if isinstance(win_start, dt.date) else pd.Timestamp(win_start).date()
    # Calibration window: normally the last 60 days (recency captures availability
    # drift). But for nodes with an EIA-923 long-history anchor, a 60-day window is
    # unreliable — at sites where ERA5 under-resolves the wind (e.g. the Rio Grande
    # Valley low-level jet) the recent window over-corrects and the factor clamps.
    # Widen those to the full available SCED record (still gen_forecast's own
    # physics, so baseline-consistent) so the factor converges to the EIA-validated
    # long-run level. Opt-in per site — unanchored nodes keep the 60-day window.
    from ercot_core import eia_anchor as _eia
    _has_anchor = bool(_eia.load(a.get("resource_node") or a.get("resource_name") or ""))
    _cal_lookback = 730 if _has_anchor else 60
    cal_start_date = max(win_start_date, win_end_date - dt.timedelta(days=_cal_lookback))

    @st.cache_data(show_spinner="Loading calibration weather (ERA5 archive)…", ttl=86400)
    def _archive(lat, lon, tech_key, start_str, end_str):
        try:
            return wf.fetch_archive(lat, lon, tech_key, start_str, end_str), None
        except Exception as exc:  # noqa: BLE001
            return None, str(exc)

    @st.cache_data(show_spinner="Calibrating against SCED history…", ttl=3600)
    def _calibrate(lat, lon, tech_key, cal_start_str, win_end_str, rnode, units_tuple,
                   cut_in_t, rated_t, cut_out_t, turbine_t, sced_units_t=()):
        arch_df, err = _archive(lat, lon, tech_key, cal_start_str, win_end_str)
        if arch_df is None:
            return 1.0, 0
        cal_start_d = dt.date.fromisoformat(cal_start_str)
        win_end_d = dt.date.fromisoformat(win_end_str)
        if cal_start_d >= win_end_d:
            return 1.0, 0
        t_start = pd.Timestamp(cal_start_d)
        t_end = pd.Timestamp(win_end_d) + pd.Timedelta(days=1)
        if units_tuple:
            gen_raw = hub.generation(rnode, list(units_tuple), t_start, t_end)
        else:
            gen_raw = hub.generation(rnode, t_start, t_end)
        if gen_raw.empty:
            return 1.0, 0
        gen_raw = gen_raw.copy()
        # Isolate this contract's resource(s) — the resource node can be SHARED by
        # multiple plants (e.g. Heart of Texas settles RTS_U1 only at node RN_RTS1,
        # which also hosts the co-located RTS2 units). Summing every unit at the
        # node would ~2× the metered output and inflate the calibration factor.
        # Mirror the settlement engine, which settles only the contract's units.
        if sced_units_t and "resource_name" in gen_raw.columns:
            _f = gen_raw[gen_raw["resource_name"].isin(sced_units_t)]
            if not _f.empty:
                gen_raw = _f
        gen_raw["mwh"] = gen_raw.get("mwh", gen_raw["mw"] * 0.25)
        gen_raw["date"] = pd.to_datetime(gen_raw["interval_start"]).dt.date
        sced_daily = gen_raw.groupby("date")["mwh"].sum() * share
        factor = gf.calibrate(arch_df, sced_daily, cap_share, tech_key, hub_height_m=hub_h,
                              cut_in=cut_in_t, rated=rated_t, cut_out=cut_out_t,
                              turbine_type=turbine_t)
        return factor, int(len(sced_daily))

    cal_factor, n_cal_days = _calibrate(
        float(a["lat"]), float(a["lon"]), tech,
        str(cal_start_date),
        str(win_end_date),
        a["resource_node"],
        tuple(_units),
        cut_in_ms, rated_ms, cut_out_ms, turbine_type,
        tuple(a.get("sced_units") or ()),
    )

    # ── bias-correct the forecast wind product to the ERA5 baseline ──────────
    # cal_factor is fit on ERA5 archive winds; the forecast API (and GEFS) run
    # hotter, so applying cal_factor to raw forecast winds over-predicts. Scale
    # the forecast winds by the ERA5/forecast mean-hub-wind ratio over their
    # recent overlap. Solar is unaffected (radiation products agree closely).
    wind_bias_ratio = None
    if tech == "wind" and not weather_df.empty:
        _ov_start = weather_df.index.min().date()
        _ov_end = today_ct.date() - dt.timedelta(days=2)   # ERA5 archive lag
        if _ov_end > _ov_start:
            _ov_arch, _ = _archive(float(a["lat"]), float(a["lon"]), tech,
                                   str(_ov_start), str(_ov_end))
            _r = _wind_bias_ratio(weather_df, _ov_arch, hub_h)
            if _r is not None:
                # Clamp rather than discard. A very hot forecast (r ≤ 0.5) is
                # exactly the case that most needs correcting — dropping it let the
                # current month over-predict (e.g. 70%+ CF). Apply a clamped ratio.
                _r = min(max(_r, 0.55), 1.45)
                wind_bias_ratio = _r
                weather_df = weather_df.copy()
                for _c in ("wind_speed_80m", "wind_speed_100m", "wind_speed_120m"):
                    if _c in weather_df.columns:
                        weather_df[_c] = pd.to_numeric(weather_df[_c], errors="coerce") * _r

    # ── prior month: ERA5 generation × actual settlement prices ──────────────
    # Shows how the model performed against real market prices last month —
    # distinct from Past Settlement which uses actual SCED generation.
    prev_month_start_dt = month_start_ct - pd.offsets.MonthBegin(1)
    prev_month_str = prev_month_start_dt.strftime("%Y-%m")
    prev_month_start_date = prev_month_start_dt.date()
    prev_month_end_date = month_start_ct.date() - dt.timedelta(days=1)

    # Resolve settlement location (node or hub) from contract terms
    _settle_loc = str(terms.get("settle_point") or "").strip()
    if not _settle_loc:
        _settle_loc = (a.get("hub", a["resource_node"])
                       if terms.get("settle_at") == "hub"
                       else a["resource_node"])
    _settle_is_hub = _settle_loc.upper().startswith("HB_")

    @st.cache_data(show_spinner=f"Loading prior-month ({prev_month_str}) ERA5 generation…", ttl=86400)
    def _prior_era5(lat, lon, tech_key, start_str, end_str):
        try:
            return wf.fetch_archive(lat, lon, tech_key, start_str, end_str), None
        except Exception as exc:  # noqa: BLE001
            return None, str(exc)

    @st.cache_data(show_spinner=f"Loading {prev_month_str} actual prices…", ttl=3600)
    def _prior_actual_prices(settle_loc, is_hub, start_ts_str, end_ts_str):
        s = pd.Timestamp(start_ts_str)
        e = pd.Timestamp(end_ts_str)
        try:
            df = hub.hub_prices(settle_loc, s, e) if is_hub else hub.node_prices(settle_loc, s, e)
            if df is None or df.empty:
                return pd.Series(dtype=float)
            df = df.copy()
            df["interval_start"] = pd.to_datetime(df["interval_start"])
            df["date"] = df["interval_start"].dt.date
            price_col = next((c for c in ("spp", "settlement_point_price") if c in df.columns), df.columns[-1])
            return df.groupby("date")[price_col].mean()
        except Exception:  # noqa: BLE001
            return pd.Series(dtype=float)

    prior_rows: list[dict] = []
    prior_net = 0.0
    prior_mwh = 0.0
    has_prior = False

    if prev_month_start_date >= win_start_date:
        _prev_era5_df, _ = _prior_era5(
            float(a["lat"]), float(a["lon"]), tech,
            str(prev_month_start_date), str(prev_month_end_date),
        )
        _prev_prices = _prior_actual_prices(
            _settle_loc, _settle_is_hub,
            str(pd.Timestamp(prev_month_start_date)),
            str(pd.Timestamp(prev_month_end_date) + pd.Timedelta(days=1)),
        )
        if _prev_era5_df is not None and not _prev_prices.empty:
            _prev_daily = gf.daily_forecast_mwh(
                _prev_era5_df, tech, cap_share,
                hub_height_m=hub_h, cal_factor=cal_factor,
                cut_in=cut_in_ms, rated=rated_ms, cut_out=cut_out_ms,
                turbine_type=turbine_type,
            )
            for _d, _mwh in _prev_daily.items():
                if _d < prev_month_start_date or _d > prev_month_end_date:
                    continue
                _price = float(_prev_prices.get(_d, float("nan")))
                if pd.isna(_price):
                    continue
                _net = float(_mwh) * (_price - strike)
                prior_rows.append({"date": _d, "mwh": float(_mwh), "net": _net,
                                   "price": _price, "kind": "retrocast"})
            prior_net = sum(r["net"] for r in prior_rows)
            prior_mwh = sum(r["mwh"] for r in prior_rows)
            has_prior = len(prior_rows) > 0

    # ── current month actuals ─────────────────────────────────────────────────
    actual_rows: list[dict] = []
    win_end_date = win_end if isinstance(win_end, dt.date) else pd.Timestamp(win_end).date()
    cur_month_start_date = month_start_ct.date()

    if win_end_date >= cur_month_start_date:
        @st.cache_data(show_spinner="Loading current-month actuals…", ttl=1800)
        def _actuals(start_str, end_str, tkey):
            s = dt.date.fromisoformat(start_str)
            e = dt.date.fromisoformat(end_str)
            res = analytics.settle(s, e, terms)
            if res is None or res["intervals"].empty:
                return pd.DataFrame()
            df = res["intervals"].copy()
            df["date"] = pd.to_datetime(df["interval_start"]).dt.date
            return df.groupby("date").agg(
                mwh=("mwh", "sum"),
                net=("cfd", "sum"),
                price=("price", "mean"),
            ).reset_index()

        da = _actuals(
            str(cur_month_start_date),
            str(min(win_end_date, today_ct.date())),
            tuple(sorted(terms.items())),
        )
        for _, row in da.iterrows():
            actual_rows.append({
                "date": row["date"], "mwh": row["mwh"],
                "net": row["net"], "price": row["price"], "kind": "actual",
            })

    settled_dates = {r["date"] for r in actual_rows}

    # ── current-month MTD fallback: ERA5 generation × actual market prices ────
    # SCED meter data publishes on a ~60-day lag, so a portal whose data lake
    # isn't gap-filled has no metered actuals for the current month and MTD
    # would read $0. Value the elapsed days exactly like the prior-month card —
    # ERA5 weather generation × actual settlement prices — so MTD is a real,
    # consistent figure for every portal. Only runs when SCED actuals are absent.
    # If the settle location is a node whose price also lags, fall back to the
    # asset's hub price (always maintained) as the market reference.
    mtd_est_rows: list[dict] = []
    mtd_price_loc = _settle_loc
    if not actual_rows:
        _mtd_end = min(today_ct.date() - dt.timedelta(days=2),
                       next_month_start.date() - dt.timedelta(days=1))
        if _mtd_end >= cur_month_start_date:
            _cur_era5_df, _ = _prior_era5(
                float(a["lat"]), float(a["lon"]), tech,
                str(cur_month_start_date), str(_mtd_end),
            )
            _mtd_is_hub = _settle_is_hub
            _cur_prices = _prior_actual_prices(
                mtd_price_loc, _mtd_is_hub,
                str(pd.Timestamp(cur_month_start_date)),
                str(pd.Timestamp(_mtd_end) + pd.Timedelta(days=1)),
            )
            if _cur_prices.empty and not _settle_is_hub:
                mtd_price_loc, _mtd_is_hub = a.get("hub", _settle_loc), True
                _cur_prices = _prior_actual_prices(
                    mtd_price_loc, _mtd_is_hub,
                    str(pd.Timestamp(cur_month_start_date)),
                    str(pd.Timestamp(_mtd_end) + pd.Timedelta(days=1)),
                )
            if _cur_era5_df is not None and not _cur_prices.empty:
                _cur_daily = gf.daily_forecast_mwh(
                    _cur_era5_df, tech, cap_share,
                    hub_height_m=hub_h, cal_factor=cal_factor,
                    cut_in=cut_in_ms, rated=rated_ms, cut_out=cut_out_ms,
                    turbine_type=turbine_type,
                )
                for _d, _mwh in _cur_daily.items():
                    if _d < cur_month_start_date or _d > _mtd_end:
                        continue
                    _price = float(_cur_prices.get(_d, float("nan")))
                    if pd.isna(_price):
                        continue
                    mtd_est_rows.append({
                        "date": _d, "mwh": float(_mwh),
                        "net": float(_mwh) * (_price - strike),
                        "price": _price, "kind": "mtd_est",
                    })
        settled_dates = settled_dates | {r["date"] for r in mtd_est_rows}

    # ── weather-forecast days ─────────────────────────────────────────────────
    daily_fcast = gf.daily_forecast_mwh(
        weather_df, tech, cap_share,
        hub_height_m=hub_h, cal_factor=cal_factor,
        cut_in=cut_in_ms, rated=rated_ms, cut_out=cut_out_ms,
        turbine_type=turbine_type,
    )
    weather_max_date = max(daily_fcast.index) if len(daily_fcast) > 0 else today_ct.date()

    # Expected market price per forward day: the month's forecast P50 when the
    # caller supplies one (so the forward price varies Jul→Oct like the projection
    # table), else the flat sidebar forward price.
    def _fpx(d) -> float:
        return float((fwd_price_by_month or {}).get(d.strftime("%Y-%m"), fwd_price))

    forecast_rows: list[dict] = []
    for d_date, mwh in daily_fcast.items():
        if d_date < cur_month_start_date or d_date in settled_dates:
            continue
        if d_date > fourth_month_end_date:
            continue
        net = float(mwh) * (_fpx(d_date) - strike)
        if d_date < next_month_start.date():
            kind = "forecast_cur"
        elif d_date < third_month_start.date():
            kind = "forecast_next"
        elif d_date < fourth_month_start.date():
            kind = "forecast_third"
        else:
            kind = "forecast_fourth"
        forecast_rows.append({"date": d_date, "mwh": float(mwh), "net": net,
                              "price": _fpx(d_date), "kind": kind})

    # Build prior-year daily MWh lookup for the climatological tail
    py_daily: dict[dt.date, float] = {}
    if py_df is not None:
        py_raw = gf.daily_forecast_mwh(py_df, tech, cap_share, hub_height_m=hub_h, cal_factor=cal_factor,
                                        cut_in=cut_in_ms, rated=rated_ms, cut_out=cut_out_ms,
                                        turbine_type=turbine_type)
        py_daily = {d_py: float(v) for d_py, v in py_raw.items()}

    # Fill days beyond the GEFS horizon — prior-year ERA5 first, flat shape as backstop
    d = weather_max_date + dt.timedelta(days=1)
    while d <= fourth_month_end_date:
        if d not in settled_dates and not any(r["date"] == d for r in forecast_rows):
            if d < next_month_start.date():
                month_tag = "cur"
            elif d < third_month_start.date():
                month_tag = "next"
            elif d < fourth_month_start.date():
                month_tag = "third"
            else:
                month_tag = "fourth"
            try:
                prior_date = d.replace(year=d.year - 1)
                py_mwh = py_daily.get(prior_date)
            except ValueError:
                py_mwh = None  # leap-day edge case
            if py_mwh is not None:
                mwh = py_mwh
                row_kind = f"clim_{month_tag}"
            else:
                mwh = gf.hist_mwh_for_date(d, hist_mwh)
                row_kind = f"hist_{month_tag}"
            forecast_rows.append({"date": d, "mwh": mwh, "net": mwh * (_fpx(d) - strike),
                                  "price": _fpx(d), "kind": row_kind})
        d += dt.timedelta(days=1)

    # ── plausibility guard: no forecast month may imply an impossible CF ───────
    # Catches both failure modes seen in the field: a hot GEFS current month
    # (~70% CF) and an under-covered far month (~1% CF). Out-of-band months are
    # rebuilt/rescaled to the historical monthly shape.
    #
    # Baseline: prefer the EIA-923 multi-year monthly shape when the node is
    # anchored (2–3 yr of independent history vs the ~1 yr of local SCED), by
    # converting the anchor's P50 capacity factors to full-month MWh at the
    # contracted share. Falls back to the SCED-derived hist_mwh otherwise.
    guard_hist = hist_mwh
    if _has_anchor and cap_share and cap_share == cap_share:
        _acf = _eia.monthly_cf_targets(
            a.get("resource_node") or a.get("resource_name") or "", "p50")
        if _acf:
            import calendar as _cal2
            guard_hist = pd.Series({
                int(m): float(cf) * cap_share * _cal2.monthrange(2025, int(m))[1] * 24.0
                for m, cf in _acf.items()})
    forecast_rows, _guard_notes = _guard_forecast_months(
        forecast_rows, hist_mwh=guard_hist, cap_share=cap_share, tech=tech,
        strike=strike, fpx=_fpx,
        blocked=settled_dates | {r["date"] for r in mtd_est_rows},
        fwin_start=cur_month_start_date, fwin_end=fourth_month_end_date,
    )
    if _guard_notes:
        _msg = ", ".join(f"{ym} ({act}, model CF {cf:.0%})" for ym, act, cf in _guard_notes)
        st.caption(f"⚠️ Forecast sanity guard adjusted {_msg} to the historical "
                   f"monthly shape — the weather model implied an out-of-range "
                   f"capacity factor for {'that month' if len(_guard_notes)==1 else 'those months'}.")

    all_rows = prior_rows + actual_rows + mtd_est_rows + forecast_rows
    if not all_rows:
        st.info("No data available for near-term projection.")
        return

    all_df = pd.DataFrame(all_rows).sort_values("date").reset_index(drop=True)

    # ── KPIs ──────────────────────────────────────────────────────────────────
    actual_mask = all_df["kind"] == "actual"
    mtd_mask = all_df["kind"].isin(["actual", "mtd_est"])
    mtd_is_est = bool(mtd_est_rows) and not actual_rows
    cur_mask = all_df["date"].apply(lambda d: str(d)[:7] == cur_month_str)
    next_mask = all_df["date"].apply(lambda d: str(d)[:7] == next_month_str)
    third_mask = all_df["date"].apply(lambda d: str(d)[:7] == third_month_str)
    fourth_mask = all_df["date"].apply(lambda d: str(d)[:7] == fourth_month_str)

    mtd_mwh = float(all_df.loc[mtd_mask, "mwh"].sum())
    mtd_net = float(all_df.loc[mtd_mask, "net"].sum())
    proj_cur = float(all_df.loc[cur_mask, "net"].sum())
    next_net = float(all_df.loc[next_mask, "net"].sum())
    next_mwh = float(all_df.loc[next_mask, "mwh"].sum())
    third_net = float(all_df.loc[third_mask, "net"].sum())
    third_mwh = float(all_df.loc[third_mask, "mwh"].sum())
    fourth_net = float(all_df.loc[fourth_mask, "net"].sum())
    fourth_mwh = float(all_df.loc[fourth_mask, "mwh"].sum())

    n_settled = int(actual_mask.sum())
    n_mtd = int(mtd_mask.sum())
    n_fcast_cur = int(all_df["kind"].str.contains("cur").sum())
    n_fcast_next = int(all_df["kind"].str.contains("next").sum())
    n_fcast_third = int(all_df["kind"].str.contains("third").sum())
    n_fcast_fourth = int(all_df["kind"].str.contains("fourth").sum())

    # Projected month-end = settled MTD + the remaining-days forecast, combined.
    remain_cur = proj_cur - mtd_net
    proj_mwh = float(all_df.loc[cur_mask, "mwh"].sum())

    # ── Row 1: primary financial cards (3 columns, more readable) ────────────
    _row1 = st.columns(3)
    if has_prior:
        _row1[0].metric(
            f"{prev_month_str} (ERA5 × actual)",
            branding.signed_money(prior_net),
            delta=f"{prior_mwh:,.0f} MWh · ERA5 generation model",
            delta_color="off",
            help=f"Prior month estimate using ERA5 weather-based generation "
                 f"× actual {'hub' if _settle_is_hub else 'node'} prices "
                 f"({_settle_loc}). Generation is modelled, not metered — "
                 f"see Past Settlement for actual SCED output.",
        )
    else:
        _row1[0].metric(
            f"Month-to-date ({cur_month_str})",
            branding.signed_money(mtd_net),
            delta=(f"{mtd_mwh:,.0f} MWh · {n_mtd} days · ERA5 × actual price"
                   if mtd_is_est else
                   f"{mtd_mwh:,.0f} MWh · {n_mtd} days settled"),
            delta_color="off",
            help=(f"SCED meter data publishes on a ~60-day lag, so the elapsed days "
                  f"of {cur_month_str} are valued with ERA5 weather generation × "
                  f"actual {mtd_price_loc} settlement prices until metered data arrives."
                  if mtd_is_est else None),
        )
    _row1[1].metric(
        "Projected month-end",
        branding.signed_money(proj_cur),
        delta=f"{proj_mwh:,.0f} MWh · {n_mtd} MTD + {n_fcast_cur} forecast days",
        delta_color="off",
        help=(f"Full-month projection = month-to-date + remaining forecast.\n\n"
              f"• Month-to-date ({n_mtd} days): {branding.signed_money(mtd_net)}\n"
              f"• Forecast ({n_fcast_cur} days): {branding.signed_money(remain_cur)}\n"
              f"• = Projected month-end: {branding.signed_money(proj_cur)}\n\n"
              f"Forecast days valued at forward price − strike: "
              f"({fwd_price:,.2f} − {strike:,.2f}) = {fwd_price - strike:,.2f} $/MWh."),
    )
    _row1[2].metric(
        "Cal. factor",
        f"{cal_factor:.3f}",
        delta=f"from {n_cal_days} SCED days" if n_cal_days else "no overlap — uncalibrated",
        delta_color="off",
        help=(
            "Weather-model output is scaled by this factor so it matches the "
            f"plant's metered SCED over the last {n_cal_days} days."
            + (f"\n\nForecast winds also bias-corrected ×{wind_bias_ratio:.3f} "
               "to the ERA5 calibration baseline (the forecast product runs "
               "hotter than ERA5). Generation uses a farm-level (multi-turbine) "
               "power curve, so daily output tracks the metered distribution "
               "instead of pegging at 100%."
               if wind_bias_ratio else
               "\n\nGeneration uses a farm-level (multi-turbine) power curve.")
        ),
    )

    # ── Row 2: monthly estimates (wider cards, easier to scan) ───────────────
    _row2_cols = 4 if has_prior else 3
    _row2 = st.columns(_row2_cols)
    _ri = 0
    if has_prior:
        _row2[_ri].metric(
            f"Month-to-date ({cur_month_str})",
            branding.signed_money(mtd_net),
            delta=(f"{mtd_mwh:,.0f} MWh · {n_mtd} days · ERA5 × actual price"
                   if mtd_is_est else
                   f"{mtd_mwh:,.0f} MWh · {n_mtd} days settled"),
            delta_color="off",
            help=(f"SCED meter data publishes on a ~60-day lag, so the elapsed days "
                  f"of {cur_month_str} are valued with ERA5 weather generation × "
                  f"actual {mtd_price_loc} settlement prices until metered data arrives."
                  if mtd_is_est else None),
        )
        _ri += 1
    _row2[_ri].metric(
        f"{next_month_str} estimate",
        branding.signed_money(next_net),
        delta=f"{next_mwh:,.0f} MWh · {n_fcast_next} days",
        delta_color="off",
    )
    _row2[_ri + 1].metric(
        f"{third_month_str} estimate",
        branding.signed_money(third_net),
        delta=f"{third_mwh:,.0f} MWh · {n_fcast_third} days",
        delta_color="off",
        help="Climatological (prior-year ERA5 / historical shape) — "
             "beyond the weather-forecast horizon, treat as indicative.",
    )
    _row2[_ri + 2].metric(
        f"{fourth_month_str} estimate",
        branding.signed_money(fourth_net),
        delta=f"{fourth_mwh:,.0f} MWh · {n_fcast_fourth} days",
        delta_color="off",
        help="Climatological (prior-year ERA5 / historical shape) — "
             "beyond the weather-forecast horizon, treat as indicative.",
    )

    # Visible calc for the projected month-end (past + forecast combined).
    st.caption(
        f"**Projected month-end ({cur_month_str})** = "
        f"month-to-date **{branding.signed_money(mtd_net)}** "
        f"({n_mtd} days, {mtd_mwh:,.0f} MWh) "
        f"+ remaining forecast **{branding.signed_money(remain_cur)}** "
        f"({n_fcast_cur} days, {proj_mwh - mtd_mwh:,.0f} MWh @ "
        f"{fwd_price:,.2f}−{strike:,.2f}={fwd_price - strike:,.2f} \\$/MWh) "
        f"= **{branding.signed_money(proj_cur)}** ({proj_mwh:,.0f} MWh)."
    )

    # ── price capture: generation-weighted vs simple mean grid price ──────────
    # Capture price = Σ(MWh × price) ÷ Σ MWh, backed out of the net column
    #   (net = MWh × (price − strike)) so it includes intraday weighting on
    #   metered days. Mean grid price = simple average of the daily settlement
    #   price. The two only diverge where prices actually vary — the realized /
    #   MTD window — since every forecast day carries one flat forward price,
    #   so the capture section is scoped to the realized + month-to-date rows.
    real_mask = all_df["kind"].isin(["retrocast", "actual", "mtd_est"])
    _real = all_df.loc[real_mask & all_df["mwh"].gt(0) & all_df["price"].notna()]
    if not _real.empty:
        _loc_lbl = "hub" if _settle_is_hub else "node"
        cap_price = float(_real["net"].sum() / _real["mwh"].sum()) + strike
        mean_price = float(_real["price"].mean())
        cap_ratio = 100.0 * cap_price / mean_price if mean_price else float("nan")
        st.markdown(
            f"**Price capture** · {_settle_loc} ({_loc_lbl}) · "
            f"{_real['date'].min()} → {_real['date'].max()} "
            f"({len(_real)} days)"
        )
        _pc = st.columns(3)
        _pc[0].metric(
            f"Capture price ({tech})",
            f"${cap_price:,.2f}/MWh",
            delta="generation-weighted",
            delta_color="off",
            help="Σ(MWh × price) ÷ Σ MWh — the settlement price the plant "
                 "actually earns, weighting each interval by how much it "
                 f"generated. Realized + month-to-date days only ({tech} "
                 "forecast days carry a flat forward price and are excluded).",
        )
        _pc[1].metric(
            "Mean grid price",
            f"${mean_price:,.2f}/MWh",
            delta=f"simple {_loc_lbl} average",
            delta_color="off",
            help=f"Time-weighted average {_loc_lbl} settlement price at "
                 f"{_settle_loc} over the same days, ignoring generation.",
        )
        _pc[2].metric(
            "Capture ratio",
            f"{cap_ratio:,.0f}%" if pd.notna(cap_ratio) else "—",
            delta="capture ÷ mean grid",
            delta_color="off",
            help="Capture price as a percent of the mean grid price. Below "
                 f"100% means {tech} output skews toward lower-priced hours; "
                 "above 100% means it correlates with higher prices.",
        )

        # ── node-vs-hub basis — the asset's true economics at its own node ──
        # From the realized capture anchor (full SCED history). Surfaces what the
        # plant earns at its node vs the hub it's measured against; for a hub-
        # settled CfD the generator bears this basis, not the offtaker.
        try:
            from ercot_core import capture_anchor
            _ca = capture_anchor.load(a.get("resource_node") or "")
        except Exception:  # noqa: BLE001 — basis is informational, never block the page
            _ca = None
        _b = (_ca or {}).get("basis") or {}
        if _b.get("basis_genweighted") is not None and _b.get("node_capture") is not None:
            _bv = float(_b["basis_genweighted"])
            _side = "below" if _bv < 0 else "above"
            _nmo = int((_ca or {}).get("n_months") or 0)
            # Honest about sample size: <12 months is a preliminary read.
            _hist = (f"{_nmo}-mo history" if _nmo else "SCED history")
            _prelim = " · ⚠ preliminary, limited history" if 0 < _nmo < 12 else ""
            _bearer = ("The CfD settles at the hub, so this basis is borne by the "
                       "generator, not the offtaker."
                       if _settle_is_hub else
                       "This contract settles at the node, so the basis flows "
                       "through to settlement.")
            st.caption(
                f"↪ **Node basis** — this plant's own node "
                f"(`{a.get('resource_node')}`) captures "
                f"**{float(_b['node_capture']):,.2f} \\$/MWh**, "
                f"**{abs(_bv):,.2f} \\$/MWh {_side}** the hub "
                f"(generation-weighted, {_hist}{_prelim}). {_bearer}"
            )

    # ── chart ─────────────────────────────────────────────────────────────────
    from plotly.subplots import make_subplots as _make_subplots

    SOLID_POS = branding.GOOD
    SOLID_NEG = branding.BAD
    FCAST_CUR_POS = "rgba(136,169,24,0.55)"
    FCAST_CUR_NEG = "rgba(178,58,72,0.50)"
    FCAST_NEXT_POS = "rgba(84,164,218,0.70)"
    FCAST_NEXT_NEG = "rgba(178,58,72,0.40)"
    FCAST_THIRD_POS = "rgba(124,99,196,0.60)"
    FCAST_THIRD_NEG = "rgba(178,58,72,0.35)"
    FCAST_FOURTH_POS = "rgba(180,140,60,0.55)"
    FCAST_FOURTH_NEG = "rgba(178,58,72,0.30)"
    RETRO_POS = "rgba(155,155,155,0.70)"
    RETRO_NEG = "rgba(178,58,72,0.55)"

    def _bar_color(row) -> str:
        pos = row["net"] >= 0
        k = row["kind"]
        if k in ("retrocast", "mtd_est"):
            return RETRO_POS if pos else RETRO_NEG
        if k == "actual":
            return SOLID_POS if pos else SOLID_NEG
        if "cur" in k:
            return FCAST_CUR_POS if pos else FCAST_CUR_NEG
        if "next" in k:
            return FCAST_NEXT_POS if pos else FCAST_NEXT_NEG
        if "third" in k:
            return FCAST_THIRD_POS if pos else FCAST_THIRD_NEG
        return FCAST_FOURTH_POS if pos else FCAST_FOURTH_NEG

    bar_colors = [_bar_color(r) for _, r in all_df.iterrows()]
    x_labels = [str(r["date"]) for _, r in all_df.iterrows()]

    # Check if we have real market prices to show in a price subplot
    real_price_kinds = {"retrocast", "actual", "mtd_est"}
    real_prices = [
        float(row["price"]) if row["kind"] in real_price_kinds and pd.notna(row["price"]) else None
        for _, row in all_df.iterrows()
    ]
    has_real_prices = any(p is not None for p in real_prices)
    # Expected (forecast) market price for the forward months — was missing, so the
    # future months showed no price line at all.
    fcast_prices = [
        float(row["price"]) if row["kind"] not in real_price_kinds and pd.notna(row["price"]) else None
        for _, row in all_df.iterrows()
    ]
    # Bridge the forecast line to the last settled price so it connects visually.
    if has_real_prices:
        _lr = max((i for i, p in enumerate(real_prices) if p is not None), default=None)
        if _lr is not None:
            fcast_prices[_lr] = real_prices[_lr]

    # Build prior-year generation lookup for the chart
    py_x, py_y = [], []
    _prior_year_label = ""
    if py_daily:
        for d_date in sorted(all_df["date"]):
            try:
                prior_date = d_date.replace(year=d_date.year - 1)
            except ValueError:
                continue
            if prior_date in py_daily:
                py_x.append(str(d_date))
                py_y.append(py_daily[prior_date])
                _prior_year_label = str(d_date.year - 1)

    # Two-row layout: main chart (settlement + gen) on top, price below
    if has_real_prices:
        fig = _make_subplots(
            rows=2, cols=1,
            shared_xaxes=True,
            row_heights=[0.75, 0.25],
            vertical_spacing=0.04,
            specs=[[{"secondary_y": True}], [{"secondary_y": False}]],
        )
    else:
        fig = _make_subplots(
            rows=1, cols=1,
            specs=[[{"secondary_y": True}]],
        )

    # ── settlement bars (row 1, primary y) ──────────────────────────────────
    if has_prior:
        fig.add_bar(x=[], y=[], name=f"ERA5 × actual – {prev_month_str}",
                    marker_color=RETRO_POS, showlegend=True, row=1, col=1)
    fig.add_bar(x=[], y=[], name="Settled", marker_color=SOLID_POS, showlegend=True, row=1, col=1)
    fig.add_bar(x=[], y=[], name=f"Forecast – {cur_month_str}", marker_color=FCAST_CUR_POS, showlegend=True, row=1, col=1)
    fig.add_bar(x=[], y=[], name=f"Forecast – {next_month_str}", marker_color=FCAST_NEXT_POS, showlegend=True, row=1, col=1)
    fig.add_bar(x=[], y=[], name=f"Forecast – {third_month_str}", marker_color=FCAST_THIRD_POS, showlegend=True, row=1, col=1)
    fig.add_bar(x=[], y=[], name=f"Forecast – {fourth_month_str}", marker_color=FCAST_FOURTH_POS, showlegend=True, row=1, col=1)

    fig.add_bar(
        x=x_labels,
        y=all_df["net"].tolist(),
        marker_color=bar_colors,
        showlegend=False,
        hovertemplate="%{x}<br>Net: $%{y:,.0f}<extra></extra>",
        row=1, col=1,
    )

    # ── generation lines (row 1, secondary y) ───────────────────────────────
    settled_mwh = [v if k == "actual" else None for v, k in zip(all_df["mwh"], all_df["kind"])]
    fcast_mwh   = [v if k != "actual" else None for v, k in zip(all_df["mwh"], all_df["kind"])]

    fig.add_scatter(
        x=x_labels, y=settled_mwh,
        mode="lines+markers",
        line=dict(color="rgba(0,105,179,0.85)", width=2),
        marker=dict(size=4),
        name="Generation (settled)",
        connectgaps=False,
        hovertemplate="%{x}<br>Gen: %{y:,.0f} MWh<extra></extra>",
        row=1, col=1, secondary_y=True,
    )
    fig.add_scatter(
        x=x_labels, y=fcast_mwh,
        mode="lines+markers",
        line=dict(color="rgba(0,105,179,0.45)", width=2, dash="dot"),
        marker=dict(size=3),
        name="Generation (forecast)",
        connectgaps=False,
        hovertemplate="%{x}<br>Gen: %{y:,.0f} MWh<extra></extra>",
        row=1, col=1, secondary_y=True,
    )

    # ── prior-year generation (row 1, secondary y, distinct orange-grey) ────
    if py_x:
        fig.add_scatter(
            x=py_x, y=py_y,
            mode="lines",
            line=dict(color="rgba(180,160,120,0.45)", width=1.5, dash="dashdot"),
            name=f"Generation ({_prior_year_label})",
            connectgaps=True,
            hovertemplate="%{x}<br>Prior yr: %{y:,.0f} MWh<extra></extra>",
            row=1, col=1, secondary_y=True,
        )

    # ── price subplot (row 2) ──────────────────────────────────────────────
    if has_real_prices:
        # Realized market price (solid)
        fig.add_scatter(
            x=x_labels, y=real_prices,
            mode="lines+markers",
            line=dict(color="rgba(218,165,32,0.85)", width=2),
            marker=dict(size=3, color="rgba(218,165,32,0.85)"),
            name="Market price ($/MWh)",
            connectgaps=True,
            hovertemplate="%{x}<br>Price: $%{y:,.2f}/MWh<extra></extra>",
            row=2, col=1,
        )

        # P10/P50/P90 forward price band from the forecast model
        _band_df = fwd_band_df
        if _band_df is not None and not _band_df.empty:
            _bx, _p10, _p50, _p90 = [], [], [], []
            for _, row in all_df.iterrows():
                mk = row["date"].strftime("%Y-%m") if hasattr(row["date"], "strftime") else str(row["date"])[:7]
                if row["kind"] in real_price_kinds:
                    continue
                brow = _band_df[_band_df["Month"] == mk]
                if brow.empty:
                    continue
                _bx.append(str(row["date"]))
                _p10.append(float(brow.iloc[0]["p10"]))
                _p50.append(float(brow.iloc[0]["p50"]))
                _p90.append(float(brow.iloc[0]["p90"]))
            if _bx:
                # Shaded P10–P90 band
                fig.add_scatter(
                    x=_bx + _bx[::-1],
                    y=_p90 + _p10[::-1],
                    fill="toself",
                    fillcolor="rgba(0,105,179,0.10)",
                    line=dict(width=0),
                    name="P10–P90 range",
                    hoverinfo="skip",
                    showlegend=True,
                    row=2, col=1,
                )
                # P50 center line
                fig.add_scatter(
                    x=_bx, y=_p50,
                    mode="lines",
                    line=dict(color="rgba(0,105,179,0.6)", width=2, dash="dot"),
                    name="P50 forecast",
                    connectgaps=True,
                    hovertemplate="%{x}<br>P50: $%{y:,.2f}/MWh<extra></extra>",
                    row=2, col=1,
                )
                # P10/P90 edges
                fig.add_scatter(
                    x=_bx, y=_p10,
                    mode="lines",
                    line=dict(color="rgba(0,105,179,0.25)", width=1, dash="dot"),
                    name="P10",
                    connectgaps=True,
                    hovertemplate="%{x}<br>P10: $%{y:,.2f}/MWh<extra></extra>",
                    showlegend=False,
                    row=2, col=1,
                )
                fig.add_scatter(
                    x=_bx, y=_p90,
                    mode="lines",
                    line=dict(color="rgba(0,105,179,0.25)", width=1, dash="dot"),
                    name="P90",
                    connectgaps=True,
                    hovertemplate="%{x}<br>P90: $%{y:,.2f}/MWh<extra></extra>",
                    showlegend=False,
                    row=2, col=1,
                )
        elif any(p is not None for p in fcast_prices):
            fig.add_scatter(
                x=x_labels, y=fcast_prices,
                mode="lines",
                line=dict(color="rgba(218,165,32,0.6)", width=2, dash="dot"),
                name="Expected market price",
                connectgaps=True,
                hovertemplate="%{x}<br>Expected: $%{y:,.2f}/MWh<extra></extra>",
                row=2, col=1,
            )

        # Strike reference line
        if x_labels:
            fig.add_scatter(
                x=[x_labels[0], x_labels[-1]], y=[strike, strike],
                mode="lines",
                line=dict(color="rgba(178,58,72,0.5)", width=1, dash="dash"),
                name=f"Strike (${strike:,.2f})",
                showlegend=True,
                hoverinfo="skip",
                row=2, col=1,
            )

    # ── shaded month backgrounds + labels ────────────────────────────────────
    bdy_str = next_month_start.strftime("%Y-%m-%d")
    bdy2_str = third_month_start.strftime("%Y-%m-%d")
    bdy3_str = fourth_month_start.strftime("%Y-%m-%d")
    cur_start_str = str(cur_month_start_date)
    fourth_end_str = str(fourth_month_end_date + dt.timedelta(days=1))

    _n_rows = 2 if has_real_prices else 1
    for _ri in range(1, _n_rows + 1):
        _yref = f"y{'' if _ri == 1 else _ri * 2 - 1}"
        if has_prior:
            prev_start_str = str(prev_month_start_date)
            fig.add_vrect(x0=prev_start_str, x1=cur_start_str,
                          fillcolor="rgba(155,155,155,0.07)", line_width=0, row=_ri, col=1)
            fig.add_vline(x=cur_start_str, line_dash="dot", line_color="#848484", line_width=1.5, row=_ri, col=1)
        fig.add_vrect(x0=cur_start_str, x1=bdy_str,
                      fillcolor="rgba(136,169,24,0.06)", line_width=0, row=_ri, col=1)
        fig.add_vrect(x0=bdy_str, x1=bdy2_str,
                      fillcolor="rgba(84,164,218,0.06)", line_width=0, row=_ri, col=1)
        fig.add_vrect(x0=bdy2_str, x1=bdy3_str,
                      fillcolor="rgba(124,99,196,0.06)", line_width=0, row=_ri, col=1)
        fig.add_vrect(x0=bdy3_str, x1=fourth_end_str,
                      fillcolor="rgba(180,140,60,0.06)", line_width=0, row=_ri, col=1)
        fig.add_vline(x=bdy_str, line_dash="dot", line_color="#848484", line_width=1.5, row=_ri, col=1)
        fig.add_vline(x=bdy2_str, line_dash="dot", line_color="#848484", line_width=1.5, row=_ri, col=1)
        fig.add_vline(x=bdy3_str, line_dash="dot", line_color="#848484", line_width=1.5, row=_ri, col=1)

    # Month summary labels (on row 1 only)
    cur_mid  = str(cur_month_start_date + dt.timedelta(days=15))
    next_mid = str(next_month_start.date() + dt.timedelta(days=15))
    third_mid = str(third_month_start.date() + dt.timedelta(days=15))
    fourth_mid = str(fourth_month_start.date() + dt.timedelta(days=15))
    _lbl = dict(showarrow=False, yref="y domain", y=1.07,
                font=dict(size=10), xanchor="center",
                bgcolor="rgba(255,255,255,0.92)",
                bordercolor="#bbb", borderwidth=1, borderpad=4)

    # Per-month average forward price for the annotation detail
    def _avg_price(mask):
        _s = all_df.loc[mask]
        if _s.empty or _s["mwh"].sum() == 0:
            return fwd_price
        return float((_s["net"].sum() / _s["mwh"].sum()) + strike)

    if has_prior:
        prev_mid = str(prev_month_start_date + dt.timedelta(days=15))
        fig.add_annotation(x=prev_mid,
                           text=f"<b>{prev_month_str}</b> ERA5×actual"
                                f"<br>{branding.signed_money_raw(prior_net)}  ·  {prior_mwh:,.0f} MWh"
                                f"<br><span style='color:#888'>{len(prior_rows)} days  ·  avg ${_avg_price(all_df['kind']=='retrocast'):,.1f}/MWh</span>",
                           **_lbl)
    fig.add_annotation(x=cur_mid,
                       text=f"<b>Current month</b> ({cur_month_str})"
                            f"<br>{branding.signed_money_raw(proj_cur)}  ·  {proj_mwh:,.0f} MWh"
                            f"<br><span style='color:#888'>{n_mtd} MTD + {n_fcast_cur} forecast days</span>",
                       **_lbl)
    fig.add_annotation(x=next_mid,
                       text=f"<b>Next month</b> ({next_month_str})"
                            f"<br>{branding.signed_money_raw(next_net)}  ·  {next_mwh:,.0f} MWh"
                            f"<br><span style='color:#888'>{n_fcast_next} days  ·  avg ${_avg_price(next_mask):,.1f}/MWh</span>",
                       **_lbl)
    fig.add_annotation(x=third_mid,
                       text=f"<b>+2 months</b> ({third_month_str})"
                            f"<br>{branding.signed_money_raw(third_net)}  ·  {third_mwh:,.0f} MWh"
                            f"<br><span style='color:#888'>{n_fcast_third} days  ·  avg ${_avg_price(third_mask):,.1f}/MWh</span>",
                       **_lbl)
    fig.add_annotation(x=fourth_mid,
                       text=f"<b>+3 months</b> ({fourth_month_str})"
                            f"<br>{branding.signed_money_raw(fourth_net)}  ·  {fourth_mwh:,.0f} MWh"
                            f"<br><span style='color:#888'>{n_fcast_fourth} days  ·  avg ${_avg_price(fourth_mask):,.1f}/MWh</span>",
                       **_lbl)

    fig.update_layout(
        height=520 if has_real_prices else 420,
        hovermode="x unified",
        margin=dict(t=30, b=10),
        legend=dict(orientation="h", y=1.14),
        bargap=0.15,
    )
    fig.update_yaxes(title_text="Daily net settlement ($)", zeroline=True, zerolinecolor="#ddd",
                     row=1, col=1, secondary_y=False)
    fig.update_yaxes(title_text="Daily generation (MWh)", showgrid=False, rangemode="tozero",
                     tickfont=dict(color="rgba(0,105,179,0.8)"),
                     title_font=dict(color="rgba(0,105,179,0.8)"),
                     row=1, col=1, secondary_y=True)
    if has_real_prices:
        fig.update_yaxes(title_text="Price ($/MWh)", showgrid=True, gridcolor="#f0f0f0",
                         tickfont=dict(color="rgba(218,165,32,0.8)"),
                         title_font=dict(color="rgba(218,165,32,0.8)"),
                         row=2, col=1)
    st.plotly_chart(fig, use_container_width=True)

    # ── detail table ──────────────────────────────────────────────────────────
    with st.expander("Daily detail"):
        show = all_df[["date", "mwh", "net", "kind"]].copy()
        show["date"] = show["date"].astype(str)
        show["mwh"] = show["mwh"].map(lambda v: f"{v:,.1f}")
        show["net"] = show["net"].map(branding.signed_money_raw)
        kind_labels = {
            "retrocast": f"ERA5 gen × actual price – {prev_month_str}",
            "actual": "Settled",
            "mtd_est": f"ERA5 gen × actual price – {cur_month_str} (MTD)",
            "forecast_cur": f"GEFS forecast – {cur_month_str}",
            "forecast_next": f"GEFS forecast – {next_month_str}",
            "clim_cur": f"Prior-year ERA5 – {cur_month_str}",
            "clim_next": f"Prior-year ERA5 – {next_month_str}",
            "hist_cur": f"Hist. shape – {cur_month_str}",
            "hist_next": f"Hist. shape – {next_month_str}",
            "forecast_third": f"GEFS forecast – {third_month_str}",
            "clim_third": f"Prior-year ERA5 – {third_month_str}",
            "hist_third": f"Hist. shape – {third_month_str}",
            "forecast_fourth": f"GEFS forecast – {fourth_month_str}",
            "clim_fourth": f"Prior-year ERA5 – {fourth_month_str}",
            "hist_fourth": f"Hist. shape – {fourth_month_str}",
        }
        show["kind"] = show["kind"].map(kind_labels).fillna(show["kind"])
        show.columns = ["Date", "MWh", "Net ($)", "Source"]
        st.dataframe(show, hide_index=True, use_container_width=True)

    # ── Excel export ─────────────────────────────────────────────────────────
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        SR_BLUE_HEX   = "0069B3"
        SR_GREEN_HEX  = "88A918"
        SR_GHOST_HEX  = "ECF0F9"
        SR_PALE_HEX   = "D7E2F2"

        kind_labels_xl = {
            "actual":       "Settled (SCED)",
            "mtd_est":      f"ERA5 gen × actual price – {cur_month_str} (MTD)",
            "forecast_cur": f"GEFS forecast – {cur_month_str}",
            "forecast_next": f"GEFS forecast – {next_month_str}",
            "clim_cur":     f"Prior-year ERA5 – {cur_month_str}",
            "clim_next":    f"Prior-year ERA5 – {next_month_str}",
            "hist_cur":     f"Hist. shape – {cur_month_str}",
            "hist_next":    f"Hist. shape – {next_month_str}",
            "forecast_third": f"GEFS forecast – {third_month_str}",
            "clim_third":   f"Prior-year ERA5 – {third_month_str}",
            "hist_third":   f"Hist. shape – {third_month_str}",
            "forecast_fourth": f"GEFS forecast – {fourth_month_str}",
            "clim_fourth":  f"Prior-year ERA5 – {fourth_month_str}",
            "hist_fourth":  f"Hist. shape – {fourth_month_str}",
        }

        def _build_excel() -> bytes:
            wb = openpyxl.Workbook()

            # ── Sheet 1: Forecast ──────────────────────────────────────────
            ws_fc = wb.active
            ws_fc.title = "Forecast"

            hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill("solid", fgColor=SR_BLUE_HEX)
            alt_fill  = PatternFill("solid", fgColor=SR_GHOST_HEX)
            act_fill  = PatternFill("solid", fgColor="E6F2E6")
            thin = Side(style="thin", color=SR_PALE_HEX)
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center")
            right  = Alignment(horizontal="right",  vertical="center")

            headers = ["Date", "MWh", "Market Price ($/MWh)",
                       "Strike ($/MWh)", "Net Settlement ($)", "Source"]
            col_widths = [14, 14, 22, 18, 22, 30]
            for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
                cell = ws_fc.cell(row=1, column=ci, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = center
                cell.border = border
                ws_fc.column_dimensions[get_column_letter(ci)].width = w
            ws_fc.row_dimensions[1].height = 20

            for ri, row in enumerate(all_df.itertuples(index=False), 2):
                is_actual = row.kind == "actual"
                fill = act_fill if is_actual else (alt_fill if ri % 2 == 0 else PatternFill())
                values = [
                    str(row.date),
                    round(float(row.mwh), 2),
                    round(float(fwd_price), 2),
                    round(float(strike), 2),
                    round(float(row.net), 2),
                    kind_labels_xl.get(row.kind, row.kind),
                ]
                aligns = [center, right, right, right, right, Alignment(vertical="center")]
                for ci, (val, aln) in enumerate(zip(values, aligns), 1):
                    cell = ws_fc.cell(row=ri, column=ci, value=val)
                    cell.alignment = aln
                    cell.fill = fill
                    cell.border = border
                    if ci in (2, 3, 4):
                        cell.number_format = '#,##0.00'
                    elif ci == 5:
                        cell.number_format = '#,##0.00'
            # Freeze header
            ws_fc.freeze_panes = "A2"

            # ── Sheet 2: Monthly summary ───────────────────────────────────
            ws_mo = wb.create_sheet("Monthly Summary")
            mo_df = all_df.copy()
            mo_df["month"] = mo_df["date"].apply(lambda d: str(d)[:7])
            mo_summary = (mo_df.groupby("month")
                          .agg(days=("date", "count"),
                               gen_mwh=("mwh", "sum"),
                               net_usd=("net", "sum"))
                          .reset_index())
            mo_summary.columns = ["Month", "Days", "Generation (MWh)", "Net Settlement ($)"]

            mo_hdr_fill = PatternFill("solid", fgColor=SR_GREEN_HEX)
            mo_headers = list(mo_summary.columns)
            mo_widths   = [12, 8, 22, 22]
            for ci, (h, w) in enumerate(zip(mo_headers, mo_widths), 1):
                cell = ws_mo.cell(row=1, column=ci, value=h)
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
                cell.fill = mo_hdr_fill
                cell.alignment = center
                cell.border = border
                ws_mo.column_dimensions[get_column_letter(ci)].width = w
            for ri, row in enumerate(mo_summary.itertuples(index=False), 2):
                vals = list(row)
                aligns2 = [center, center, right, right]
                for ci, (val, aln) in enumerate(zip(vals, aligns2), 1):
                    cell = ws_mo.cell(row=ri, column=ci, value=val)
                    cell.alignment = aln
                    cell.border = border
                    if ci in (3, 4):
                        cell.number_format = '#,##0.00'
            ws_mo.freeze_panes = "A2"

            # ── Sheet 3: About ─────────────────────────────────────────────
            ws_ab = wb.create_sheet("About")
            about_rows = [
                ("Asset",           str(a.get("project_name", a.get("resource_node", "")))),
                ("Resource node",   str(a.get("resource_node", ""))),
                ("Technology",      str(a.get("tech", ""))),
                ("Capacity (MW)",   float(a.get("capacity_mw", 0))),
                ("Volume share",    f"{share*100:.4g}%"),
                ("Strike ($/MWh)",  float(strike)),
                ("Fwd price ($/MWh)", float(fwd_price)),
                ("Cal. factor",     round(float(cal_factor), 4)),
                ("Cal. days (SCED)", int(n_cal_days)),
                ("Current month",   cur_month_str),
                ("Next month",      next_month_str),
                ("Projected cur. ($)", round(float(proj_cur), 2)),
                ("Estimated next ($)", round(float(next_net), 2)),
                ("Generated",       str(pd.Timestamp.now("America/Chicago").strftime("%Y-%m-%d %H:%M CT"))),
                ("Source",          "Open-Meteo (16-day) + GEFS P50 (35-day) + ERA5 climatology"),
            ]
            ab_key_font = Font(name="Calibri", bold=True, color=SR_BLUE_HEX, size=11)
            ws_ab.column_dimensions["A"].width = 26
            ws_ab.column_dimensions["B"].width = 42
            for ri, (key, val) in enumerate(about_rows, 1):
                ck = ws_ab.cell(row=ri, column=1, value=key)
                cv = ws_ab.cell(row=ri, column=2, value=val)
                ck.font = ab_key_font
                ck.fill = PatternFill("solid", fgColor=SR_GHOST_HEX)
                ck.border = border
                cv.border = border

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        xl_bytes = _build_excel()
        asset_slug = str(a.get("project_name", "forecast")).lower().replace(" ", "_")
        fname = f"{asset_slug}_forecast_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"
        st.download_button(
            "⬇ Download forecast Excel",
            data=xl_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Daily generation forecast + monthly summary + metadata",
        )
    except ImportError:
        pass  # openpyxl not installed in this venv

    # ── footnote ─────────────────────────────────────────────────────────────
    src = "Open-Meteo free API (no key required)"
    py_note = (
        f"Beyond 35 days: prior-year ERA5 (same calendar days, {month_start_ct.year - 1})."
        if py_daily else "Beyond 35 days: historical monthly shape."
    )
    st.caption(
        f"**Source:** {src}. "
        f"**Cal. factor {cal_factor:.3f}** — weather-model output scaled to match "
        f"{'last ' + str(n_cal_days) + ' days of SCED history' if n_cal_days else 'no SCED history (uncalibrated)'}. "
        f"Near-term: high-res forecast (16 days) then GEFS ensemble P50 (35 days). {py_note} "
        f"Forward price: **\\${fwd_price:,.2f}/MWh** · Strike: **\\${strike:,.2f}/MWh**."
    )

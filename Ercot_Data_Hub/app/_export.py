"""Unified data export for the Streamlit app — CSV, Excel, Markdown, PDF.

One entry point, :func:`download_block`, renders a compact format picker + a
single download button next to any table. Bytes for the *selected* format only
are built per run (cheap for large frames), and any format whose optional
dependency is missing is dropped gracefully rather than erroring.

Excel and PDF lead with a **Summary** sheet / cover page that actually explains
the export to a non-technical reader: a titled header, when it was generated,
the report details (``meta``), plain-English ``notes``, and a **column guide**
(auto-built from the data, enriched by an optional ``glossary``). The aim is
that someone opening the file with no other context understands what they're
looking at and what every column means.

Engines: CSV/Markdown via pandas (+tabulate), Excel via openpyxl, PDF via fpdf2.
"""

from __future__ import annotations

import io

import pandas as pd

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Caps so a huge frame can't blow up a PDF (paginated tables stay legible) or an
# over-wide PDF spill off the page; CSV/Excel/Markdown are uncapped.
_PDF_MAX_ROWS = 500
_PDF_MAX_COLS = 14
_MD_MAX_ROWS = 5000

# SR Inc. palette (hex without '#', for openpyxl / fpdf).
_NAVY = "1F3A5F"
_LIGHT = "EAF0F6"
_GREY = "6B7682"


def _stamp() -> str:
    # Date.now is fine here (display/file-name only; never persisted state).
    return pd.Timestamp.now().strftime("%Y%m%d-%H%M")


def _stamp_human() -> str:
    return pd.Timestamp.now().strftime("%B %d, %Y at %H:%M")


def _meta_lines(meta: dict | None) -> list[str]:
    return [f"{k}: {v}" for k, v in (meta or {}).items()]


def _friendly_dtype(s: pd.Series) -> str:
    """A non-technical description of a column's data type."""
    if pd.api.types.is_datetime64_any_dtype(s):
        return "date/time"
    if pd.api.types.is_bool_dtype(s):
        return "yes/no"
    if pd.api.types.is_integer_dtype(s):
        return "whole number"
    if pd.api.types.is_float_dtype(s):
        return "number"
    return "text"


def _example(s: pd.Series) -> str:
    """A representative non-null value, formatted for display."""
    nn = s.dropna()
    if nn.empty:
        return ""
    v = nn.iloc[0]
    if isinstance(v, float):
        return f"{v:,.4g}"
    return str(v)[:40]


def _column_guide(df: pd.DataFrame, glossary: dict | None) -> pd.DataFrame:
    """One row per column: name, plain-English meaning, type, fill rate, example."""
    glossary = {str(k).lower(): v for k, v in (glossary or {}).items()}
    n = len(df)
    rows = []
    for c in df.columns:
        s = df[c]
        rows.append({
            "Column": str(c),
            "What it means": glossary.get(str(c).lower(), ""),
            "Type": _friendly_dtype(s),
            "Filled": f"{int(s.notna().sum()):,} / {n:,}",
            "Example": _example(s),
        })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------- #
# Per-format encoders
# --------------------------------------------------------------------------- #

def to_csv_bytes(df: pd.DataFrame, **_) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


# ── Excel ──────────────────────────────────────────────────────────────────
def to_excel_bytes(df: pd.DataFrame, *, sheet_name: str = "Data",
                   title: str | None = None, meta: dict | None = None,
                   notes: list[str] | None = None,
                   glossary: dict | None = None, **_) -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    data_sheet = "Data"
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name=data_sheet, index=False)
        wb = xl.book
        ws = wb.create_sheet("Summary", 0)        # cover sheet first
        _excel_summary(ws, df, title, meta, notes, glossary, data_sheet,
                       Alignment, Font, PatternFill, get_column_letter)
        _excel_style_data(wb[data_sheet], df, Alignment, Font, PatternFill,
                          get_column_letter)
        wb.active = 0
    return buf.getvalue()


def _excel_summary(ws, df, title, meta, notes, glossary, data_sheet,
                   Alignment, Font, PatternFill, get_column_letter):
    navy = PatternFill("solid", fgColor=_NAVY)
    light = PatternFill("solid", fgColor=_LIGHT)
    white_b = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    grey_i = Font(color=_GREY, italic=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 30

    r = 1
    # Title banner
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, _xl(title or "Data export"))
    c.font = Font(color="FFFFFF", bold=True, size=15)
    c.fill = navy
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 26
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    sub = ws.cell(r, 1, f"Generated {_stamp_human()}  ·  {len(df):,} rows × {df.shape[1]} columns")
    sub.font = grey_i
    sub.alignment = Alignment(indent=1)
    r += 2

    # Report details (meta)
    if meta:
        ws.cell(r, 1, "REPORT DETAILS").font = white_b
        ws.cell(r, 1).fill = navy
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for k, v in meta.items():
            ws.cell(r, 1, _xl(k)).font = bold
            ws.cell(r, 1).fill = light
            vc = ws.cell(r, 2, _xl(v))
            vc.alignment = wrap
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            r += 1
        r += 1

    # Notes
    if notes:
        ws.cell(r, 1, "NOTES").font = white_b
        ws.cell(r, 1).fill = navy
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for note in notes:
            cell = ws.cell(r, 1, "•  " + _xl(note))
            cell.alignment = wrap
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.row_dimensions[r].height = max(15, 15 * (1 + len(str(note)) // 90))
            r += 1
        r += 1

    # Column guide
    guide = _column_guide(df, glossary)
    ws.cell(r, 1, f"COLUMN GUIDE — the '{data_sheet}' sheet").font = white_b
    ws.cell(r, 1).fill = navy
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    headers = list(guide.columns)
    for j, h in enumerate(headers, 1):
        hc = ws.cell(r, j, h)
        hc.font = bold
        hc.fill = light
    r += 1
    for _, row in guide.iterrows():
        for j, h in enumerate(headers, 1):
            cell = ws.cell(r, j, _xl(row[h]))
            cell.alignment = wrap
        r += 1
    ws.freeze_panes = "A4"


def _excel_style_data(ws, df, Alignment, Font, PatternFill, get_column_letter):
    """Bold/filled header, frozen top row, autofilter, sensible widths."""
    navy = PatternFill("solid", fgColor=_NAVY)
    head = Font(color="FFFFFF", bold=True)
    for j, col in enumerate(df.columns, 1):
        cell = ws.cell(1, j)
        cell.font = head
        cell.fill = navy
        cell.alignment = Alignment(vertical="center")
        # width from header + a sample of values (Python strs — robust to Arrow NA)
        lengths = sorted(len(str(v)) for v in df[col].head(200).tolist())
        p90 = lengths[int(len(lengths) * 0.9)] if lengths else 0
        width = max(len(str(col)) + 2, p90 + 2)
        ws.column_dimensions[get_column_letter(j)].width = min(max(width, 10), 42)
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = f"A1:{get_column_letter(df.shape[1])}{len(df) + 1}"


def _xl(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v)


# ── Markdown ─────────────────────────────────────────────────────────────────
def to_markdown_bytes(df: pd.DataFrame, *, title: str | None = None,
                     meta: dict | None = None, notes: list[str] | None = None,
                     glossary: dict | None = None, **_) -> bytes:
    out: list[str] = []
    if title:
        out += [f"# {title}", "", f"_Generated {_stamp_human()} · {len(df):,} rows × "
                f"{df.shape[1]} columns_", ""]
    if meta:
        out.append("## Report details\n")
        out += [f"- **{k}:** {v}" for k, v in meta.items()]
        out.append("")
    if notes:
        out.append("## Notes\n")
        out += [f"- {n}" for n in notes]
        out.append("")
    out.append("## Column guide\n")
    out.append(_column_guide(df, glossary).to_markdown(index=False))
    out += ["", "## Data\n"]
    body, note = df, ""
    if len(df) > _MD_MAX_ROWS:
        body = df.head(_MD_MAX_ROWS)
        note = f"\n\n_Showing first {_MD_MAX_ROWS:,} of {len(df):,} rows._"
    out.append(body.to_markdown(index=False))
    return ("\n".join(out) + note).encode("utf-8")


# ── PDF ──────────────────────────────────────────────────────────────────────
# fpdf2's built-in Helvetica is latin-1 only; map common unicode to ASCII and
# replace anything else so a stray "×"/"≈"/emoji in the data can't crash the PDF.
_PDF_SUBS = str.maketrans({"—": "-", "–": "-", "…": "...", "×": "x", "≈": "~",
                           "•": "-", "→": "->", "≥": ">=", "≤": "<=", "’": "'",
                           "“": '"', "”": '"', " ": " ", "\xa0": " ", "§": "Sec.",
                           "✅": "[OK]", "⚠": "[!]", "️": "", "📦": "", "🔍": "",
                           "−": "-"})


def _pdf_safe(s) -> str:
    return str(s).translate(_PDF_SUBS).encode("latin-1", "replace").decode("latin-1")


def _pdf_cell(x) -> str:
    if pd.isna(x):
        return ""
    s = f"{x:,.2f}" if isinstance(x, float) else str(x)
    s = s if len(s) <= 30 else s[:29] + "..."
    return _pdf_safe(s)


def _hex_rgb(h: str):
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def to_pdf_bytes(df: pd.DataFrame, *, title: str | None = None,
                meta: dict | None = None, notes: list[str] | None = None,
                glossary: dict | None = None, **_) -> bytes:
    from fpdf import FPDF  # optional dep; caller catches ImportError

    clipped_rows = len(df) > _PDF_MAX_ROWS
    clipped_cols = df.shape[1] > _PDF_MAX_COLS
    d = df.head(_PDF_MAX_ROWS).iloc[:, :_PDF_MAX_COLS]

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    navy = _hex_rgb(_NAVY)

    # ── cover page (portrait-feel header, full-width sections) ──
    pdf.add_page()
    pdf.set_fill_color(*navy)
    pdf.rect(0, 0, pdf.w, 18, style="F")
    pdf.set_xy(10, 4)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 10, _pdf_safe(title or "Data export"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(90, 90, 90)
    pdf.set_xy(10, 20)
    pdf.set_font("Helvetica", "I", 9)
    pdf.cell(0, 6, _pdf_safe(f"Generated {_stamp_human()}  -  {len(df):,} rows x "
                             f"{df.shape[1]} columns"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    def section(label):
        pdf.set_fill_color(*navy)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 7, "  " + _pdf_safe(label), new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(1)

    if meta:
        section("Report details")
        pdf.set_font("Helvetica", size=9)
        for k, v in meta.items():
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(60, 5, _pdf_safe(f"{k}:"))
            pdf.set_font("Helvetica", size=9)
            pdf.multi_cell(0, 5, _pdf_safe(v), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    if notes:
        section("Notes")
        pdf.set_font("Helvetica", size=9)
        for n in notes:
            pdf.multi_cell(0, 5, "- " + _pdf_safe(n), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    # Column guide
    section("Column guide")
    guide = _column_guide(df, glossary)[["Column", "What it means", "Type", "Example"]]
    pdf.set_font("Helvetica", size=8)
    with pdf.table(line_height=4.4, cell_fill_color=235, cell_fill_mode="ROWS",
                   first_row_as_headings=True,
                   col_widths=(28, 52, 14, 22)) as table:
        hr = table.row()
        for h in guide.columns:
            hr.cell(_pdf_safe(h))
        for _, row in guide.iterrows():
            tr = table.row()
            for v in row:
                tr.cell(_pdf_safe(v))

    # ── data table (new page) ──
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Data", new_x="LMARGIN", new_y="NEXT")
    headers = [_pdf_safe(c) for c in d.columns]
    rows = [[_pdf_cell(v) for v in r] for r in d.itertuples(index=False, name=None)]
    pdf.set_font("Helvetica", size=7)
    with pdf.table(line_height=4.2, cell_fill_color=235, cell_fill_mode="ROWS",
                   first_row_as_headings=True) as table:
        hr = table.row()
        for h in headers:
            hr.cell(h)
        for r in rows:
            tr = table.row()
            for cv in r:
                tr.cell(cv)

    if clipped_rows or clipped_cols:
        pdf.ln(1)
        pdf.set_font("Helvetica", "I", 7)
        bits = []
        if clipped_rows:
            bits.append(f"first {_PDF_MAX_ROWS:,} of {len(df):,} rows")
        if clipped_cols:
            bits.append(f"first {_PDF_MAX_COLS} of {df.shape[1]} columns")
        pdf.cell(0, 4, "PDF shows " + " and ".join(bits)
                 + " - use CSV/Excel for the full table.", new_x="LMARGIN", new_y="NEXT")
    return bytes(pdf.output())


# format key -> (encoder, extension, mime)
_FORMATS = {
    "CSV": (to_csv_bytes, "csv", "text/csv"),
    "Excel": (to_excel_bytes, "xlsx", EXCEL_MIME),
    "Markdown": (to_markdown_bytes, "md", "text/markdown"),
    "PDF": (to_pdf_bytes, "pdf", "application/pdf"),
}


def _available(formats) -> list[str]:
    out = []
    for f in formats:
        if f == "Excel":
            try:
                import openpyxl  # noqa: F401
            except Exception:  # noqa: BLE001
                continue
        if f == "Markdown":
            try:
                import tabulate  # noqa: F401
            except Exception:  # noqa: BLE001
                continue
        if f == "PDF":
            try:
                import fpdf  # noqa: F401
            except Exception:  # noqa: BLE001
                continue
        out.append(f)
    return out


# --------------------------------------------------------------------------- #
# UI
# --------------------------------------------------------------------------- #

def download_block(st, df: pd.DataFrame, *, name: str, title: str | None = None,
                   meta: dict | None = None, notes: list[str] | None = None,
                   glossary: dict | None = None,
                   formats=("CSV", "Excel", "Markdown", "PDF"),
                   container=None, key: str | None = None) -> None:
    """Render a format picker + download button for ``df``.

    ``name`` seeds the download file name. ``title``/``meta`` head the Summary
    cover sheet (Excel) / cover page (PDF) and the Markdown header. ``notes`` is
    a list of plain-English explanations shown on that cover. ``glossary`` maps
    column name → meaning for the auto-built **column guide** (columns without an
    entry still appear, just without a description). Only the chosen format's
    bytes are built per run; formats whose optional dependency is missing are
    silently dropped.
    """
    target = container or st
    if df is None or len(df) == 0:
        target.caption("Nothing to export yet.")
        return

    avail = _available(formats)
    if not avail:
        return
    key = key or name
    # No st.columns here — keep it nestable inside callers' column layouts.
    fmt = target.radio("⬇ Export as", avail, horizontal=True, key=f"{key}_fmt")
    encoder, ext, mime = _FORMATS[fmt]
    try:
        data = encoder(df, title=title, meta=meta, notes=notes, glossary=glossary,
                       sheet_name=name)
    except Exception as exc:  # noqa: BLE001 — never let an export crash the page
        target.warning(f"{fmt} export unavailable: {exc}")
        return
    target.download_button(f"⬇ Download {fmt}", data,
                           file_name=f"{name}_{_stamp()}.{ext}", mime=mime,
                           key=f"{key}_dl")
